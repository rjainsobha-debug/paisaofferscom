from __future__ import annotations

import json
import re
import time
from collections import Counter
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import requests
from openpyxl import load_workbook

try:
    from bs4 import BeautifulSoup
except ModuleNotFoundError:
    BeautifulSoup = None


ROOT = Path(__file__).resolve().parents[1]
SHEET_NAME = "SiteStripe Queue"
EXCEL_CANDIDATES = [
    ROOT / "tools" / "amazon_deal_queue.xlsx",
    ROOT / "data" / "exports" / "amazon_deal_queue.xlsx",
]
OUTPUT_JSON = ROOT / "data" / "deals.json"
REPORT_PATH = ROOT / "tools" / "publish_report.txt"
IMAGE_CACHE_PATH = ROOT / "tools" / "image_cache.json"
FALLBACK_IMAGE = "/opengraph.jpg"
REQUEST_TIMEOUT_SECONDS = 10
FETCH_DELAY_SECONDS = 0.75
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36 PaisaOffersPublisher/1.0"
)
ALLOWED_LINK_HOSTS = ("amazon.in", "amazon.com", "amzn.in", "amzn.to")
PRICE_TEXT_FIELDS = ("clean_title", "short_message", "notes")


class ImageResolver:
    def __init__(self, cache_path: Path):
        self.cache_path = cache_path
        self.cache: dict[str, dict[str, str]] = self._load_cache()
        self.images_from_excel = 0
        self.images_fetched = 0
        self.images_from_cache = 0
        self.image_fetch_failed = 0
        self.fallback_images_used = 0
        self._last_fetch_at = 0.0

    def resolve(self, asin: str, product_url: str) -> str:
        asin = (asin or "").strip().upper()
        cached = self.cache.get(asin)
        if cached and cached.get("image_url"):
            cached_image_url = str(cached.get("image_url") or "").strip()
            if _is_publishable_image_url(cached_image_url) and cached.get("source") != "fallback":
                self.images_from_cache += 1
                return cached_image_url

        image_url = ""
        source = ""
        try:
            image_url, source = self._fetch_image(asin, product_url)
        except Exception:
            image_url = ""

        if image_url:
            self.images_fetched += 1
            self.cache[asin] = {
                "asin": asin,
                "image_url": image_url,
                "fetched_at": _now_iso(),
                "source": source or "amazon_page",
            }
            return image_url

        self.image_fetch_failed += 1
        return ""

    def save(self) -> None:
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        self.cache_path.write_text(json.dumps(self.cache, indent=2, ensure_ascii=False), encoding="utf-8")

    def _fetch_image(self, asin: str, product_url: str) -> tuple[str, str]:
        url = product_url or f"https://www.amazon.in/dp/{asin}"
        self._polite_delay()
        response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=REQUEST_TIMEOUT_SECONDS)
        response.raise_for_status()
        html = response.text or ""
        if _looks_blocked(html):
            return "", ""
        if BeautifulSoup is None:
            return _image_from_html_without_bs4(html, url)
        soup = BeautifulSoup(html, "html.parser")

        og_image = _clean_image_url((soup.find("meta", property="og:image") or {}).get("content"), url)
        if og_image:
            return og_image, "og:image"

        landing_image = soup.find("img", id="landingImage")
        if landing_image:
            src_image = _clean_image_url(landing_image.get("src"), url)
            if src_image:
                return src_image, "landingImage:src"
            dynamic_image = _dynamic_image_url(landing_image.get("data-a-dynamic-image"), url)
            if dynamic_image:
                return dynamic_image, "landingImage:data-a-dynamic-image"
        return "", ""

    def _polite_delay(self) -> None:
        elapsed = time.monotonic() - self._last_fetch_at
        if elapsed < FETCH_DELAY_SECONDS:
            time.sleep(FETCH_DELAY_SECONDS - elapsed)
        self._last_fetch_at = time.monotonic()

    def _load_cache(self) -> dict[str, dict[str, str]]:
        if not self.cache_path.exists():
            return {}
        try:
            data = json.loads(self.cache_path.read_text(encoding="utf-8"))
            return data if isinstance(data, dict) else {}
        except (OSError, json.JSONDecodeError):
            return {}


def main() -> int:
    excel_path = _find_excel_file()
    if not excel_path:
        print("ERROR: Could not find tools/amazon_deal_queue.xlsx or data/exports/amazon_deal_queue.xlsx")
        return 1

    rows = _read_rows(excel_path)
    image_resolver = ImageResolver(IMAGE_CACHE_PATH)
    skipped = Counter()
    approved_rows = []
    candidates_by_asin: dict[str, dict[str, Any]] = {}

    for row in rows:
        if str(row.get("review_status", "")).strip().upper() != "APPROVED":
            skipped["not_approved"] += 1
            continue
        approved_rows.append(row)

        reason = _validation_skip_reason(row)
        if reason:
            skipped[reason] += 1
            continue

        asin = str(row.get("asin", "")).strip().upper()
        existing = candidates_by_asin.get(asin)
        if existing is None or _row_sort_date(row) > _row_sort_date(existing):
            candidates_by_asin[asin] = row

    deals = []
    for row in candidates_by_asin.values():
        asin = str(row.get("asin", "")).strip().upper()
        normalized_url = str(row.get("normalized_amazon_url", "")).strip() or f"https://www.amazon.in/dp/{asin}"
        image_url = _product_image_url_from_row(row)
        if image_url:
            image_resolver.images_from_excel += 1
        else:
            image_url = image_resolver.resolve(asin, normalized_url)
        if not image_url:
            skipped["missing_product_image"] += 1
            continue
        deal = _deal_from_row(row, image_url)
        deals.append(deal)

    image_resolver.save()
    deals.sort(key=lambda item: _parse_date(item.get("date_added")), reverse=True)
    hot_deals = sorted(deals, key=lambda item: _number(item.get("discount_percent")), reverse=True)[:6]
    for item in hot_deals:
        item["is_hot_deal"] = True

    payload = {
        "generated_at": _now_iso(),
        "total_deals": len(deals),
        "hot_deals": hot_deals,
        "all_deals": deals,
    }
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    published_count = len(deals)
    skipped_count = sum(skipped.values())
    _write_report(
        excel_path=excel_path,
        total_rows=len(rows),
        approved_rows=len(approved_rows),
        published_count=published_count,
        skipped_count=skipped_count,
        skipped=skipped,
        image_resolver=image_resolver,
    )
    _print_summary(len(rows), len(approved_rows), published_count, skipped_count, skipped, image_resolver)
    return 0


def _find_excel_file() -> Path | None:
    for path in EXCEL_CANDIDATES:
        if path.exists():
            return path
    return None


def _read_rows(excel_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise SystemExit(f"ERROR: Missing required sheet: {SHEET_NAME}")
        sheet = workbook[SHEET_NAME]
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        return rows
    finally:
        workbook.close()


def _validation_skip_reason(row: dict[str, Any]) -> str:
    _populate_missing_prices(row)
    required = {
        "asin": row.get("asin"),
        "deal_price": row.get("deal_price"),
        "site_stripe_affiliate_link": row.get("site_stripe_affiliate_link"),
    }
    for field, value in required.items():
        if _is_missing(value):
            return f"missing_{field}"
    if not _is_allowed_amazon_link(str(row.get("site_stripe_affiliate_link", ""))):
        return "invalid_affiliate_link"
    if not str(row.get("clean_title", "")).strip():
        return "missing_clean_title"
    if _number(row.get("deal_price")) <= 0:
        return "invalid_deal_price"
    return ""


def _deal_from_row(row: dict[str, Any], image_url: str) -> dict[str, Any]:
    source_name = str(row.get("source_name", "") or "").strip()
    return {
        "asin": str(row.get("asin", "") or "").strip().upper(),
        "title": str(row.get("clean_title", "") or "").strip(),
        "image_url": image_url,
        "deal_price": _number(row.get("deal_price")),
        "original_price": _number(row.get("original_price")),
        "discount_percent": _number(row.get("discount_pct")),
        "category": str(row.get("category", "") or source_name).strip(),
        "affiliate_link": str(row.get("site_stripe_affiliate_link", "") or "").strip(),
        "description": str(row.get("short_message", "") or "").strip(),
        "date_added": _date_text(row.get("collected_at")),
        "expiry_date": str(row.get("expiry_date", "") or "").strip(),
        "tags": [],
        "is_hot_deal": False,
    }


def _is_allowed_amazon_link(url: str) -> bool:
    text = (url or "").strip()
    if not text:
        return False
    parsed = urlparse(text if re.match(r"^[a-z][a-z0-9+.-]*://", text, re.I) else f"https://{text}")
    host = (parsed.netloc or parsed.path.split("/", 1)[0]).lower()
    if "@" in host:
        host = host.rsplit("@", 1)[-1]
    host = host.split(":", 1)[0].strip(".")
    return any(host == allowed or host.endswith(f".{allowed}") for allowed in ALLOWED_LINK_HOSTS)


def _populate_missing_prices(row: dict[str, Any]) -> None:
    if _is_missing(row.get("deal_price")):
        extracted_price = _extract_first_from_text(row, extract_price_from_text)
        if extracted_price > 0:
            row["deal_price"] = extracted_price

    if _is_missing(row.get("discount_pct")):
        extracted_discount = _extract_first_from_text(row, extract_discount_from_text)
        if extracted_discount > 0:
            row["discount_pct"] = extracted_discount

    deal_price = _number(row.get("deal_price"))
    discount_pct = _number(row.get("discount_pct"))
    if _is_missing(row.get("original_price")) and deal_price > 0 and 1 <= discount_pct <= 95:
        row["original_price"] = round(deal_price / (1 - discount_pct / 100), 2)

    if _is_missing(row.get("original_price")) and deal_price > 0:
        row["original_price"] = deal_price


def _extract_first_from_text(row: dict[str, Any], extractor) -> float:
    for field in PRICE_TEXT_FIELDS:
        value = extractor(row.get(field))
        if value > 0:
            return value
    return 0.0


def extract_price_from_text(text: Any) -> float:
    value = str(text or "")
    if not value.strip():
        return 0.0
    patterns = (
        r"\bprice\s*:\s*(?:rs\.?\s*)?(?:inr\s*)?(?:₹\s*)?(\d[\d,]*(?:\.\d+)?)\s*(?:rs\.?|inr|₹)?",
        r"\bstarting\s*@\s*(?:rs\.?\s*)?(?:inr\s*)?(?:₹\s*)?(\d[\d,]*(?:\.\d+)?)",
        r"\bjust\s*(?:rs\.?\s*)?(?:inr\s*)?₹?\s*(\d[\d,]*(?:\.\d+)?)",
        r"@\s*(?:rs\.?\s*)?(?:inr\s*)?₹\s*(\d[\d,]*(?:\.\d+)?)",
        r"\brs\.?\s*(\d[\d,]*(?:\.\d+)?)",
        r"₹\s*(\d[\d,]*(?:\.\d+)?)",
    )
    for pattern in patterns:
        match = re.search(pattern, value, re.I)
        if match:
            return float(match.group(1).replace(",", ""))
    return 0.0


def extract_discount_from_text(text: Any) -> float:
    value = str(text or "")
    if not value.strip():
        return 0.0
    match = re.search(r"\bdiscount\s*:\s*(\d+(?:\.\d+)?)\s*%\s*off\b", value, re.I)
    if not match:
        match = re.search(r"\b(\d+(?:\.\d+)?)\s*%\s*off\b", value, re.I)
    return float(match.group(1)) if match else 0.0


def _row_sort_date(row: dict[str, Any]) -> datetime:
    return max(_parse_date(row.get("reviewed_at")), _parse_date(row.get("collected_at")))


def _parse_date(value: Any) -> datetime:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)
    text = str(value or "").strip()
    if not text:
        return datetime.min.replace(tzinfo=timezone.utc)
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).astimezone(timezone.utc)
    except ValueError:
        return datetime.min.replace(tzinfo=timezone.utc)


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    return str(value or "").strip()


def _number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "")
    match = re.search(r"-?\d+(?:,\d{2,3})*(?:\.\d+)?|-?\d+(?:\.\d+)?", text)
    return float(match.group(0).replace(",", "")) if match else 0.0


def _is_missing(value: Any) -> bool:
    if value in (None, ""):
        return True
    if isinstance(value, float) and value != value:
        return True
    return not str(value).strip()


def _product_image_url_from_row(row: dict[str, Any]) -> str:
    value = str(row.get("product_image_url", "") or "").strip()
    return value if _is_publishable_image_url(value) else ""


def _is_publishable_image_url(value: str) -> bool:
    text = str(value or "").strip()
    lowered = text.lower()
    return lowered.startswith("http") and not lowered.endswith(FALLBACK_IMAGE)


def _clean_image_url(value: str | None, base_url: str) -> str:
    if not value:
        return ""
    value = unescape(str(value)).strip().strip("\"'")
    if value.startswith("//"):
        value = f"https:{value}"
    if value.startswith("/"):
        value = urljoin(base_url, value)
    if not value.startswith(("http://", "https://")):
        return ""
    return value.split()[0]


def _dynamic_image_url(raw_value: str | None, base_url: str) -> str:
    if not raw_value:
        return ""
    try:
        data = json.loads(unescape(raw_value))
    except json.JSONDecodeError:
        return ""
    if not isinstance(data, dict):
        return ""
    for url in data:
        cleaned = _clean_image_url(url, base_url)
        if cleaned:
            return cleaned
    return ""


def _image_from_html_without_bs4(html: str, base_url: str) -> tuple[str, str]:
    og_match = re.search(
        r"""<meta[^>]+property=["']og:image["'][^>]+content=["']([^"']+)["']""",
        html,
        re.I,
    ) or re.search(
        r"""<meta[^>]+content=["']([^"']+)["'][^>]+property=["']og:image["']""",
        html,
        re.I,
    )
    if og_match:
        image_url = _clean_image_url(og_match.group(1), base_url)
        if image_url:
            return image_url, "og:image"

    img_match = re.search(r"""<img[^>]+id=["']landingImage["'][^>]+src=["']([^"']+)["']""", html, re.I)
    if img_match:
        image_url = _clean_image_url(img_match.group(1), base_url)
        if image_url:
            return image_url, "landingImage:src"
    return "", ""


def _looks_blocked(html: str) -> bool:
    lowered = (html or "").lower()
    return "captcha" in lowered or "robot check" in lowered or "enter the characters you see below" in lowered


def _now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def _write_report(
    *,
    excel_path: Path,
    total_rows: int,
    approved_rows: int,
    published_count: int,
    skipped_count: int,
    skipped: Counter,
    image_resolver: ImageResolver,
) -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"excel_path: {excel_path}",
        f"sheet_name: {SHEET_NAME}",
        f"total rows read: {total_rows}",
        f"approved rows found: {approved_rows}",
        f"published count: {published_count}",
        f"skipped count: {skipped_count}",
        "skipped reasons:",
    ]
    if skipped:
        lines.extend(f"  {reason}: {count}" for reason, count in sorted(skipped.items()))
    else:
        lines.append("  none: 0")
    lines.extend(
        [
            f"images_from_excel: {image_resolver.images_from_excel}",
            f"images_from_cache: {image_resolver.images_from_cache}",
            f"images_fetched: {image_resolver.images_fetched}",
            f"image_fetch_failed: {image_resolver.image_fetch_failed}",
            f"missing_product_image: {skipped.get('missing_product_image', 0)}",
            f"fallback_images_used: {image_resolver.fallback_images_used}",
        ]
    )
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _print_summary(
    total_rows: int,
    approved_rows: int,
    published_count: int,
    skipped_count: int,
    skipped: Counter,
    image_resolver: ImageResolver,
) -> None:
    print("PaisaOffers publish summary")
    print(f"Rows read: {total_rows}")
    print(f"Approved rows found: {approved_rows}")
    print(f"Published deals: {published_count}")
    print(f"Skipped rows: {skipped_count}")
    if skipped:
        print("Skipped reasons:")
        for reason, count in sorted(skipped.items()):
            print(f"  {reason}: {count}")
    print(f"images_from_excel: {image_resolver.images_from_excel}")
    print(f"images_from_cache: {image_resolver.images_from_cache}")
    print(f"images_fetched: {image_resolver.images_fetched}")
    print(f"image_fetch_failed: {image_resolver.image_fetch_failed}")
    print(f"missing_product_image: {skipped.get('missing_product_image', 0)}")
    print(f"fallback_images_used: {image_resolver.fallback_images_used}")
    print(f"Wrote: {OUTPUT_JSON}")
    print(f"Report: {REPORT_PATH}")


if __name__ == "__main__":
    raise SystemExit(main())
